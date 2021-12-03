from openpyxl.reader.excel import load_workbook
from pywebio.input import *
from pywebio.output import *
from pywebio.pin import *
from pywebio import start_server
import os
from fpdf import FPDF
import glob
import pandas as pd
from openpyxl import Workbook
from datetime import datetime
import csv

os.system('cls')

def app():
    
    popup('Warning!!', [
        put_text("Clicking the following Delete button will delete all your files in the transcriptIITP folder. Otherwise you can Close the popup and move on"),
        put_buttons(['Delete','Close'], onclick=[lambda: clear_folder(), lambda: close_popup()])
    ])


    # input form
    info = input_group("Upload CSV files", [
        file_upload("Upload Roll Number Sheet", name="file1", required=True, accept=".csv"),
        file_upload("Upload Subject Master Sheet",  name="file2", required=True, accept=".csv"),
        file_upload("Upload Grades Sheet",  name="file3", required=True, accept=".csv"),
        file_upload("Upload SEAL", name="img1", accept="image/*", multiple=False, required=False),
        file_upload("Upload Signature", name="img2", accept="image/*", multiple=False, required=False)
    ])

    file1 = info['file1']
    file2 = info['file2']
    file3 = info['file3']
    img1 = info['img1']
    img2 = info['img2']

    # saving names-roll.csv
    content1 = file1['content'].decode('utf-8').splitlines()
    save_csv(content1, "names-roll")

    # saving subjects_master.csv 
    content2 = file2['content'].decode('utf-8').splitlines()
    save_csv(content2, "subjects_master")

    # saving grades.csv 
    content3 = file3['content'].decode('utf-8').splitlines()
    save_csv(content3, "grades")

    # saving both seal and signature
    if img1:
        img1["filename"] = "SEAL."+img1["filename"].split('.')[1]
        open(r'input/'+img1["filename"], 'wb').write(img1['content'])
    else:
        for name in glob.glob("input/SEAL*"):
            os.remove(name)
    
    if img2:
        img2["filename"] = "Signature."+img2["filename"].split('.')[1]
        open(r'input/'+img2["filename"], 'wb').write(img2['content']) 
    else:
        for name in glob.glob("input/Signature*"):
            os.remove(name)

    workfunc()


def workfunc():
    clear()

    # Format required by the user
    put_input("range", label="Type the range of Roll Numbers")
    put_button("Generate Range Transcripts", onclick=lambda:generate_transcripts(pin["range"]), color='success', outline=True)
    put_button("Generate All Transcripts", onclick=lambda: generate_transcripts("0000AA00-9999ZZ99"), color='success', outline=True)


def clear_folder():
    for filename in glob.glob("transcriptIITP/*"):
        os.remove(filename)
    close_popup()


def progress():
    put_html("<p align=""center"">\
    <img src=""https://media0.giphy.com/media/kUTME7ABmhYg5J3psM/200.webp?\
    cid=ecf05e47som5hu3l2owou9vmn20hue70j113dgls1ghb1909&rid=200.webp&ct=g""\
    width=""120px""></p>")


def generate_transcripts(rolln):
    clear()
    progress()
    with open('input/names-roll.csv', 'r') as csvfile:
        csv_reader = csv.DictReader(csvfile)
        names = {}

        for row in csv_reader:
            names[row['Roll'].upper()] = row['Name']

    with open('input/grades.csv', 'r') as csvfile:
        csv_reader = csv.DictReader(csvfile)
        record_grades = {}

        for row in csv_reader:
            listA = []
            listA.append(row['Roll'])
            listA.append(row['Sem'])
            listA.append(row['SubCode'])
            listA.append(row['Credit'])
            listA.append(row['Grade'])
            listA.append(row['Sub_Type'])

            if(record_grades.get(row['Roll']) == None):
                record_grades[row['Roll'].upper()] = [listA]
            else:
                record_grades[row['Roll'].upper()].append(listA)

    with open('input/subjects_master.csv', 'r') as csvfile:
        csvreader = csv.DictReader(csvfile)
        subjects = {}

        for row in csvreader:
            listB = []
            listB.append(row['subname'])
            listB.append(row['ltp'])
            subjects[row['subno']] = listB

    results = {}
    for roll in record_grades:
        studentResult = {}

        for row in record_grades[roll]:
            sub_type = row[5]
            grade = row[4]
            credit = row[3]
            subno = row[2]
            sem = row[1]
            
            listC = []
            listC.append(subno)
            listC.append(subjects[subno][0])
            listC.append(subjects[subno][1])
            listC.append(credit)
            listC.append(sub_type)
            listC.append(grade)

            if(studentResult.get(sem) == None):
                studentResult[sem] = [listC]
            else:
                studentResult[sem].append(listC)

        results[roll] = studentResult

    st = rolln.split("-")[0].upper()
    en = rolln.split("-")[1].upper()

    for roll in record_grades:
        if roll < st or roll > en:
            continue

        if os.path.exists("output/"+roll+'.pdf'):
            continue

        if roll[2]+roll[3]=='01':
            pdf = FPDF('L', 'mm', 'A3')
            pdf.add_page()
            pdf.set_font('Arial',size= 12)
            pdf.y = 10
            
            start_y=pdf.y
            pdf.image('iitp_logo_1.jpeg', x = None, y = None, w = 0, h = 0, type = 'jpeg')
            image_bottom_y=pdf.y

            pdf.y = pdf.y+5
            pdf.x= pdf.x + 90
            pdf.cell(225,10,"     Roll No:  %s                                        Name:  %s                    Year of Admission:  %s" %(roll,names[roll],'20'+roll[:2]),border="L,R,T")
            pdf.ln(5)
            pdf.x= pdf.x + 90
            pdf.cell(225,10,"     Programe:  Bachelor of Technology               Course:  %s" %roll[4:6],border="L,R,B")
            pdf.ln(10)
            
            creds_taken = []
            totalcreds_taken = []
            sem = []
            spi = []
            totalcreds_sum = 0
            cpi = []
            cpi_sum = 0

            for i in results[roll]:
                sem.append(i)
                spi_sum = 0
                cred_sum = 0

                for row in results[roll][i]:
                    cred = float(row[3])
                    marks = float(grade_to_marks(row[5]))
                    
                    spi_sum += marks*cred
                    cred_sum += cred

                totalcreds_sum += cred_sum
                cpi_sum += (spi_sum/cred_sum)*cred_sum

                spi.append(round(spi_sum/cred_sum, 2))
                cpi.append(round(cpi_sum/totalcreds_sum, 2))
                creds_taken.append(cred_sum)
                totalcreds_taken.append(totalcreds_sum)

            pdf.ln(10)
            top_y = pdf.y
            top1_x = pdf.x
            top_x = pdf.x
            max_y = 0

            for i in results[roll]:
                if i == '10':
                    break

                if int(i) == 5:
                    pdf.line(top1_x-3,max_y+5,pdf.x+4,max_y+5)
                    pdf.y = max_y +5
                    pdf.x = top1_x 
                    top_x = top1_x 
                    top_y = max_y + 5
                    

                pdf.set_font('Arial',style="BU",size= 10)
                pdf.cell(100,10,"Semester"+i,align='L',ln=2)
                
                pdf.set_font('Arial',size= 6)
                offset_x = pdf.x

                pdf.cell(20,5,"Sub. Code",ln=0,align='C',border= 1)
                pdf.cell(40,5,"Subject Name",ln=0,align='C',border= 1)
                pdf.cell(15,5,"L-T-P",ln=0,align='C',border= 1)
                pdf.cell(10,5,"CRD",ln=0,align='C',border= 1)
                pdf.cell(10,5,"GRD",ln=1,align='C',border= 1)
                
                pdf.x= offset_x
                for row in results[roll][i]:
                    offset_x = pdf.x
                    pdf.cell(20,5,row[0],ln=0,align='C',border= 1)
                    pdf.cell(40,5,row[1],ln=0,align='C',border= 1)
                    pdf.cell(15,5,row[2],ln=0,align='C',border= 1)
                    pdf.cell(10,5,row[3],ln=0,align='C',border= 1)
                    pdf.cell(10,5,row[5],ln=1,align='C',border= 1)
                    pdf.x= offset_x

                pdf.y = pdf.y+5
                
                pdf.cell(70,7,txt="Credits Taken:  %s   Credits Cleared: %s  CPI:  %s   SPI:  %s" %(creds_taken[int(i)-1],creds_taken[int(i)-1],cpi[int(i)-1],spi[int(i)-1]), border=1,ln=1,align="C")

                max_y = max(max_y, pdf.y)       
                top_x = top_x+ 100
                pdf.x = top_x
                pdf.y = top_y

            pdf.line(top1_x-3, max_y + 5,top_x+4, max_y+5)
                

            pdf.x = top1_x + 5
            pdf.y = max_y + 25

            pdf.set_font('Arial',style="B",size= 10)    
            pdf.cell(30,7,"Date Generated: ",border=0,align='L')
            pdf.set_font('Arial',style="U",size= 10)
            today = datetime.today()
            pdf.cell(50,7,today.strftime("%d %b %Y %H:%M:%S"),border=0,align="C")

            pdf.x = pdf.x + 115
            pdf.y = pdf.y - 10

            vis = False
            for filename in glob.glob("input/SEAL*"):
                vis = True
                file = filename

            if vis:
                pdf.image(file,w = 35,h = 20)

            pdf.x = pdf.x + 150
            pdf.y = max_y + 30
            pdf.set_font('Arial',style="B",size= 10)
            
            vis = False
            for filename in glob.glob("input/Signature*"):
                vis = True
                file = filename    
            
            pdf.cell(50,7,"Assistant Registrar(Academic)",border='T',align='L')
            pdf.y = pdf.y - 22
            pdf.x = pdf.x - 45
            
            if vis:
                pdf.image(file,w = 35,h = 20)
            else:
                pdf.y = pdf.y + 20
                pdf.x = pdf.x + 35

            pdf.line(top1_x-3, pdf.y+10,top_x+4, pdf.y+10)
            pdf.line(top1_x-3,start_y,top_x+4,start_y)
            pdf.line(top1_x-3,image_bottom_y,top_x+4,image_bottom_y)
            pdf.line(top1_x-3, pdf.y+10,top1_x-3,start_y)
            pdf.line(top_x+4, start_y,top_x+4, pdf.y+10)


            path = 'TranscriptIITP/' + roll + '.pdf'
            pdf.output(path, 'F')
        
        else:
            pdf = FPDF('P', 'mm', 'A4')
            pdf.add_page()
            pdf.set_font('Arial',size= 8)
            pdf.y = 10
            
            start_y=pdf.y
            pdf.image('iitp_logo_1.jpeg', x = None, y = None, w = 190, h = 0, type = 'jpeg')
            image_bottom_y=pdf.y

            pdf.y = pdf.y+5
            pdf.x= pdf.x + 30
            pdf.cell(130,10,"  Roll No:  %s                                  Name:  %s                  Year of Admission:  %s" %(roll,names[roll],'20'+roll[:2]),border="L,R,T")
            pdf.ln(5)
            pdf.x= pdf.x + 30
            if roll[2]+roll[3] == "11":
                pdf.cell(130,10,"  Programe:  Master of Technology            Course:  %s" %roll[4:6],border="L,R,B")
            elif roll[2]+roll[3] == "12":
                pdf.cell(130,10,"  Programe:  Master of Science            Course:  %s" %roll[4:6],border="L,R,B")
            else:
                pdf.cell(130,10,"  Programe:  Doctor of Philosophy            Course:  %s" %roll[4:6],border="L,R,B")
            pdf.ln(5)
            
            creds_taken = []
            totalcreds_taken = []
            sem = []
            spi = []
            totalcreds_sum = 0
            cpi = []
            cpi_sum = 0

            for i in results[roll]:
                sem.append(i)
                spi_sum = 0
                cred_sum = 0

                for row in results[roll][i]:
                    cred = float(row[3])
                    marks = float(grade_to_marks(row[5]))
                    
                    spi_sum += marks*cred
                    cred_sum += cred

                totalcreds_sum += cred_sum
                cpi_sum += (spi_sum/cred_sum)*cred_sum

                spi.append(round(spi_sum/cred_sum, 2))
                cpi.append(round(cpi_sum/totalcreds_sum, 2))
                creds_taken.append(cred_sum)
                totalcreds_taken.append(totalcreds_sum)

            pdf.ln(5)
            top_y = pdf.y
            top1_x = pdf.x
            top_x = pdf.x
            max_y = 0
            max_x = 200

            for i in results[roll]:
                if i == '10':
                    break

                if int(i) == 3:
                    pdf.line(top1_x-3,max_y+5,pdf.x+4,max_y+5)
                    pdf.y = max_y +5
                    pdf.x = top1_x 
                    top_x = top1_x 
                    top_y = max_y + 5
                    

                pdf.set_font('Arial',style="BU",size= 9)
                pdf.cell(95,10,"Semester"+i,align='L',ln=2)
                
                pdf.set_font('Arial',size= 5)
                offset_x = pdf.x

                pdf.cell(10,5,"Sub. Code",ln=0,align='C',border= 1)
                pdf.cell(55,5,"Subject Name",ln=0,align='C',border= 1)
                pdf.cell(10,5,"L-T-P",ln=0,align='C',border= 1)
                pdf.cell(5,5,"CRD",ln=0,align='C',border= 1)
                pdf.cell(10,5,"GRD",ln=1,align='C',border= 1)
                
                pdf.x= offset_x
                for row in results[roll][i]:
                    offset_x = pdf.x
                    pdf.cell(10,5,row[0],ln=0,align='C',border= 1)
                    pdf.cell(55,5,row[1],ln=0,align='C',border= 1)
                    pdf.cell(10,5,row[2],ln=0,align='C',border= 1)
                    pdf.cell(5,5,row[3],ln=0,align='C',border= 1)
                    pdf.cell(10,5,row[5],ln=1,align='C',border= 1)
                    pdf.x= offset_x

                pdf.y = pdf.y+5
                
                pdf.cell(70,7,txt="Credits Taken:  %s   Credits Cleared: %s  CPI:  %s   SPI:  %s" %(creds_taken[int(i)-1],creds_taken[int(i)-1],cpi[int(i)-1],spi[int(i)-1]), border=1,ln=1,align="C")

                max_y = max(max_y, pdf.y)       
                top_x = top_x + 95
                pdf.x = top_x
                pdf.y = top_y

            pdf.line(top1_x-3, max_y + 5,max_x+4, max_y+5)
                

            pdf.x = top1_x + 5
            pdf.y = max_y + 25

            pdf.set_font('Arial',style="B",size= 8)    
            pdf.cell(30,7,"Date Generated: ",border=0,align='L')
            pdf.set_font('Arial',style="U",size= 8)
            today = datetime.today()
            pdf.cell(20,7,today.strftime("%d %b %Y %H:%M:%S"),border=0,align="C")

            pdf.x = pdf.x + 20
            pdf.y = pdf.y - 10

            vis = False
            for filename in glob.glob("input/SEAL*"):
                vis = True
                file = filename

            if vis:
                pdf.image(file,w = 35,h = 20)

            pdf.x = pdf.x + 60
            pdf.y = max_y + 30
            pdf.set_font('Arial',style="B",size= 8)  
            pdf.cell(30,7,"Assistant Registrar(Academic)",border='T',align='C')
            pdf.y = pdf.y - 20
            pdf.x = pdf.x - 35

            vis = False
            for filename in glob.glob("input/Signature*"):
                vis = True
                file = filename

            if vis:
                pdf.image(file,w = 35,h = 20)
            else:
                pdf.y = pdf.y + 20
                pdf.x = pdf.x + 35

            pdf.line(top1_x-3, pdf.y+10,max_x+4, pdf.y+10)
            pdf.line(top1_x-3,start_y,max_x+4,start_y)
            pdf.line(top1_x-3,image_bottom_y,max_x+4,image_bottom_y)
            pdf.line(top1_x-3, pdf.y+10,top1_x-3,start_y)
            pdf.line(max_x+4, start_y,max_x+4, pdf.y+10)


            path = 'TranscriptIITP/' + roll + '.pdf'
            pdf.output(path, 'F')

    workfunc()


def grade_to_marks(grade):
    if 'AA' in grade:
        return 10
    elif 'AB' in grade:
        return 9
    elif 'BB' in grade:
        return 8
    elif 'BC' in grade:
        return 7
    elif 'CC' in grade:
        return 6
    elif 'CD' in grade:
        return 5
    elif 'DD' in grade:
        return 4
    elif 'F' in grade or 'I' in grade:
        return 0


# function to save the content of the uploaded csv by the filename provided
def save_csv(content: list, filename):
    if os.path.exists(r"input//")==False:
        os.mkdir(r"input//")

    with open(r"input//" + '%s.csv' %filename, "w+") as csv_file:
        for line in content:
            csv_file.write(line + "\n")


if __name__=='__main__':
    start_server(app, port=36536, debug=True)
